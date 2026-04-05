"""
استيراد بيانات ملف الإكسيل إلى قاعدة البيانات SQLite
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from datetime import datetime, date
from database import SessionLocal, create_tables
from database import (Setting, Item, Supplier, Customer, Employee, Payroll,
                      Purchase, Sale, Expense, Collection, PriceBoard,
                      Receivable, Custody, ExpenseCategory)

EXCEL_PATH = r"d:\Libyca\منظومة ليبيكا الحسابية.xlsx"

CATS = {1:'نقل',2:'ديزل/وقود',3:'عمالة',4:'صيانة',5:'كهرباء',
        6:'مياه',7:'إيجار',8:'اتصالات',9:'مصاريف إدارية',10:'أخرى',
        11:'مقدم بضاعة',12:'شوالات',13:'مطبخ'}

def to_date(val):
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date):     return val
    return None

def safe_float(val, default=0.0):
    try:    return float(val) if val is not None else default
    except: return default

def safe_str(val):
    return str(val).strip() if val is not None else ""

def safe_int(val):
    try:    return int(val) if val is not None else None
    except: return None


def import_all():
    create_tables()
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    db = SessionLocal()

    # ── الإعدادات ──────────────────────────────────────────
    print("استيراد الإعدادات...")
    ws = wb['الإعدادات']
    settings_map = {
        'اسم الشركة':'company_name', 'العملة الأساسية':'currency',
        'مجال النشاط':'business_type', 'واتساب ١':'whatsapp1', 'واتساب ٢':'whatsapp2',
    }
    collected = {}
    for row in ws.iter_rows(values_only=True):
        for i in range(len(row) - 1):
            if row[i] in settings_map and row[i+1] is not None:
                key = settings_map[row[i]]
                if key not in collected:
                    collected[key] = safe_str(row[i+1])
    for key, value in collected.items():
        db.add(Setting(key=key, value=value))
    db.commit()
    print(f"  تم. ({len(collected)} إعداد)")

    # ── تصنيفات المصروفات ──────────────────────────────────
    print("استيراد تصنيفات المصروفات...")
    for cid, cname in CATS.items():
        if not db.query(ExpenseCategory).filter_by(id=cid).first():
            db.add(ExpenseCategory(id=cid, name=cname))
    db.commit()

    # ── الأصناف ────────────────────────────────────────────
    print("استيراد الأصناف...")
    ws = wb['الأصناف']
    items_dict = {}   # code -> name  (للبحث لاحقاً)
    count = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        code = safe_int(row[0])
        name = safe_str(row[1])
        if not code or not name:
            continue
        code_str = str(code)
        items_dict[code] = name
        if not db.query(Item).filter_by(code=code_str).first():
            db.add(Item(
                code=code_str, name=name,
                category=safe_str(row[2]),
                active=True if row[3] in [1,'1',True,'نعم'] else False,
                notes=safe_str(row[4]),
                avg_cost=safe_float(row[5])
            ))
            count += 1
    db.commit()
    print(f"  تم استيراد {count} صنف. items_dict={items_dict}")

    # ── الموردون ───────────────────────────────────────────
    print("استيراد الموردون...")
    ws = wb['الموردون']
    count = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        code = safe_str(row[0]); name = safe_str(row[1])
        if not name: continue
        if not db.query(Supplier).filter_by(code=code).first():
            db.add(Supplier(
                code=code, name=name,
                phone=safe_str(row[2]), city=safe_str(row[3]),
                active=True if row[4] in [1,'1',True,'نعم'] else False,
                notes=safe_str(row[5])
            ))
            count += 1
    db.commit()
    print(f"  تم استيراد {count} مورد.")

    # ── العملاء ────────────────────────────────────────────
    print("استيراد العملاء...")
    ws = wb['العملاء']
    count = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        code = safe_str(row[0]); name = safe_str(row[1])
        if not name: continue
        if not db.query(Customer).filter_by(code=code).first():
            db.add(Customer(
                code=code, name=name,
                phone=safe_str(row[2]), city=safe_str(row[3]),
                active=True if row[4] in [1,'1',True,'نعم'] else False,
                notes=safe_str(row[5])
            ))
            count += 1
    db.commit()
    print(f"  تم استيراد {count} عميل.")

    # ── الموظفون ───────────────────────────────────────────
    print("استيراد الموظفون...")
    ws = wb['الموظفون']
    count = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        code = safe_str(row[0]); name = safe_str(row[1])
        if not name: continue
        if not db.query(Employee).filter_by(code=code).first():
            db.add(Employee(
                code=code, name=name,
                department=safe_str(row[2]),
                hire_date=to_date(row[3]),
                basic_salary=safe_float(row[4]),
                shift_count=int(safe_float(row[5])),
                shift_salary=safe_float(row[6]),
                fixed_deductions=safe_float(row[7]),
                payment_method=safe_str(row[8]),
                active=True if row[9] in [1,'1',True,'نعم'] else False,
            ))
            count += 1
    db.commit()
    print(f"  تم استيراد {count} موظف.")

    # ── الرواتب ────────────────────────────────────────────
    # الرواتب فارغة في الملف الحالي - سيتم إدخالها يدوياً من النظام

    # ── المشتريات ──────────────────────────────────────────
    print("استيراد المشتريات...")
    ws = wb['المشتريات']
    count = 0
    seq = 1
    for row in ws.iter_rows(min_row=3, values_only=True):
        dt = to_date(row[1])
        if not dt:
            continue

        item_code = safe_int(row[3])
        item_name = safe_str(row[4]) or items_dict.get(item_code, '')
        supplier_name = safe_str(row[2])

        gross_weight = safe_float(row[6])
        tare_weight  = safe_float(row[7])
        net_weight   = gross_weight - tare_weight  # محسوب
        unit_price   = safe_float(row[10])
        amount_lyd   = round(net_weight * unit_price, 3)  # محسوب

        # المدفوع والمتبقي: إذا كان مُدخل نستخدمه، وإلا نفترض مدفوع بالكامل
        paid      = safe_float(row[15], default=None)
        remaining = safe_float(row[16], default=None)
        if paid is None and amount_lyd > 0:
            paid = amount_lyd
            remaining = 0.0
        elif paid is not None and remaining is None:
            remaining = amount_lyd - paid

        ref = safe_str(row[0]) or f"ش{seq:06d}"
        seq += 1

        db.add(Purchase(
            ref_no=ref, date=dt,
            supplier_name=supplier_name,
            item_code=str(item_code) if item_code else '',
            item_name=item_name,
            gross_weight=gross_weight,
            tare_weight=tare_weight,
            net_weight=net_weight,
            unit_price=unit_price,
            amount=amount_lyd,
            currency="د.ل", exchange_rate=1,
            amount_lyd=amount_lyd,
            payment_method=safe_str(row[14]),
            paid=paid or 0,
            remaining=remaining or 0,
            notes=safe_str(row[17]),
        ))
        count += 1
    db.commit()
    print(f"  تم استيراد {count} عملية شراء.")

    # حساب متوسط تكلفة كل صنف بعد الاستيراد
    print("  حساب متوسطات التكلفة...")
    from sqlalchemy import func
    for code, name in items_dict.items():
        code_str = str(code)
        avg = db.query(func.avg(Purchase.unit_price)).filter(
            Purchase.item_code == code_str,
            Purchase.unit_price > 0
        ).scalar() or 0
        item = db.query(Item).filter_by(code=code_str).first()
        if item and avg:
            item.avg_cost = round(avg, 3)
    db.commit()

    # ── المبيعات ───────────────────────────────────────────
    print("استيراد المبيعات...")
    ws = wb['المبيعات']
    count = 0
    seq = 1
    # أعد قراءة متوسط التكلفة لكل صنف
    avg_costs = {}
    for code in items_dict:
        code_str = str(code)
        avg = db.query(func.avg(Purchase.unit_price)).filter(
            Purchase.item_code == code_str, Purchase.unit_price > 0
        ).scalar() or 0
        avg_costs[code] = round(avg, 3)

    for row in ws.iter_rows(min_row=3, values_only=True):
        dt = to_date(row[1])
        if not dt:
            continue

        item_code    = safe_int(row[3])
        item_name    = safe_str(row[4]) or items_dict.get(item_code, '')
        customer_name= safe_str(row[2])
        quantity     = safe_float(row[5])
        unit_price   = safe_float(row[6])
        amount_lyd   = round(quantity * unit_price, 2)   # محسوب
        avg_cost     = avg_costs.get(item_code, 0)
        cogs         = round(quantity * avg_cost, 2)      # محسوب
        gross_profit = round(amount_lyd - cogs, 2)        # محسوب

        paid      = safe_float(row[11], default=None)
        remaining = safe_float(row[12], default=None)
        if paid is None and amount_lyd > 0:
            paid = amount_lyd; remaining = 0.0
        elif paid is not None and remaining is None:
            remaining = amount_lyd - paid

        ref = safe_str(row[0]) or f"ب{seq:06d}"
        seq += 1

        db.add(Sale(
            ref_no=ref, date=dt,
            customer_name=customer_name,
            item_code=str(item_code) if item_code else '',
            item_name=item_name,
            quantity=quantity,
            unit_price=unit_price,
            amount=amount_lyd,
            currency="د.ل", exchange_rate=1,
            amount_lyd=amount_lyd,
            cogs=cogs,
            gross_profit=gross_profit,
            payment_method=safe_str(row[10]),
            paid=paid or 0,
            remaining=remaining or 0,
            notes=safe_str(row[16]),
        ))
        count += 1
    db.commit()
    print(f"  تم استيراد {count} عملية بيع.")

    # ── المصروفات ──────────────────────────────────────────
    print("استيراد المصروفات...")
    ws = wb['المصروفات']
    count = 0
    seq = 1
    for row in ws.iter_rows(min_row=3, values_only=True):
        dt = to_date(row[1])
        description = safe_str(row[3])
        if not dt or not description:
            continue
        cat_id   = int(safe_float(row[2])) if row[2] else 10
        cat_name = CATS.get(cat_id, 'أخرى')
        amount   = safe_float(row[4])

        ref = safe_str(row[0]) or f"م{seq:06d}"
        seq += 1

        db.add(Expense(
            ref_no=ref, date=dt,
            category_id=cat_id, category_name=cat_name,
            description=description,
            amount=amount, currency="د.ل", exchange_rate=1,
            amount_lyd=amount,
            payment_method=safe_str(row[8]),
            notes=safe_str(row[10]),
        ))
        count += 1
    db.commit()
    print(f"  تم استيراد {count} مصروف.")

    # ── بورصة الأسعار ──────────────────────────────────────
    print("استيراد بورصة الأسعار...")
    ws = wb['بورصة_الأسعار']
    count = 0
    default_date = date(2026, 3, 1)
    for row in ws.iter_rows(min_row=3, values_only=True):
        item_code = safe_int(row[1])
        buy_price = safe_float(row[3])
        if not item_code or not buy_price:
            continue
        item_name = items_dict.get(item_code, f'صنف {item_code}')
        dt = to_date(row[0]) or default_date
        db.add(PriceBoard(
            date=dt,
            item_code=str(item_code),
            item_name=item_name,
            buy_price=buy_price,
            sell_price=safe_float(row[4]),
            currency=safe_str(row[5]) or "د.ل",
            exchange_rate=safe_float(row[6]) or 1,
        ))
        count += 1
    db.commit()
    print(f"  تم استيراد {count} سعر.")

    # ── العهدة ─────────────────────────────────────────────
    print("استيراد العهدة...")
    ws = wb['العهدة']
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        dt = to_date(row[1])
        amount = safe_float(row[3])
        if not dt or not amount:
            continue
        db.add(Custody(
            description=safe_str(row[0]) or "استلام عهدة",
            date=dt, amount=amount,
        ))
        count += 1
    db.commit()
    print(f"  تم استيراد {count} سجل عهدة.")

    db.close()
    print("\n✅ اكتمل الاستيراد بنجاح!")


if __name__ == "__main__":
    import_all()
